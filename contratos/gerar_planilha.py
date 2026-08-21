# -*- coding: utf-8 -*-
"""
Gera a planilha consolidada dos contratos com markup aplicado sobre todos os valores.

Estrutura do arquivo gerado:
  Parâmetros         - percentuais de markup (célula editável) e legenda de cores
  Resumo             - um contrato por linha, com totais originais e com markup
  Itens Consolidados - tabela mestre com todos os itens de todos os contratos
  CT-xxx             - uma aba por contrato, com cabeçalho, itens e subtotais
  Conferência        - confronto do calculado x valores impressos nos PDFs

Regra do markup: aplicado sobre o VALOR UNITÁRIO de mão de obra; o total do item
é recalculado como quantidade x valor unitário majorado. Como a quantidade não muda,
o total majorado equivale ao total original x (1 + markup).

Uso:  python3 gerar_planilha.py
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D

from dados_contratos import CONTRATOS

ARQUIVO_SAIDA = "Contratos_Recanto_Cataratas_com_markup.xlsx"
# Logo oficial da empresa. Se o arquivo existir nesta pasta, é embutido na faixa
# azul do topo de todas as abas; se não existir, a faixa é gerada sem o logo.
ARQUIVO_LOGO = "logo.png"

# --------------------------------------------------------------------------
# Estilo
# --------------------------------------------------------------------------
FONTE = "Arial"

# Paleta da marca Martins Notari, tirada do logo:
# grafite do fundo, amarelo do monograma e branco da tipografia.
GRAFITE = "2E2E2E"        # grafite institucional (faixa do topo, totais)
GRAFITE_CLARO = "4A4A4A"  # grafite secundário (cabeçalhos de tabela)
AMARELO = "FFC91E"        # amarelo da marca (destaques e células editáveis)
AMARELO_ESCURO = "E0A800" # amarelo de apoio
AMARELO_SUAVE = "FFF8E1"  # amarelo diluído (linhas de aditivo)
CINZA_CLARO = "F4F5F7"
CINZA_MEDIO = "E8E8E8"
CINZA_LINHA = "C7C7C7"
VERDE_CLARO = "E7F5EC"
LARANJA_CLARO = AMARELO_SUAVE  # itens de aditivo
TINTA = "2E2E2E"
TINTA_FRACA = "6B7280"

# nomes mantidos para não espalhar renomeações pelo arquivo
AZUL_ESCURO = GRAFITE
AZUL_MEDIO = GRAFITE_CLARO

COR_ENTRADA = "1155CC"   # azul: valor digitado (vem do PDF)
COR_FORMULA = "2E2E2E"    # grafite: fórmula na própria aba
COR_LINK = "15803D"       # verde: link para outra aba

CUR = 'R$ #,##0.00'
CUR4 = 'R$ #,##0.0000'
QTY = '#,##0.0000'
PCT = '0.0%'

fina = Side(style="thin", color="BFBFBF")
BORDA = Border(left=fina, right=fina, top=fina, bottom=fina)


def f(bold=False, size=10, color="000000", italic=False):
    return Font(name=FONTE, bold=bold, size=size, color=color, italic=italic)


def estilo_cabecalho(ws, linha, col_ini, col_fim, fill=AZUL_MEDIO):
    for c in range(col_ini, col_fim + 1):
        cel = ws.cell(row=linha, column=c)
        cel.font = f(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor=fill)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA


# Faixa institucional: fundo azul da marca, logo à esquerda e título recuado.
# O recuo é dado em caracteres (indent), por isso funciona em qualquer aba,
# independentemente da largura das colunas.
RECUO_LOGO = 15


def titulo(ws, texto, col_fim, subtitulo=None):
    for c in range(1, col_fim + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=AZUL_ESCURO)

    ws["A1"] = texto
    ws["A1"].font = f(bold=True, size=14, color="FFFFFF")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_fim)
    ws["A1"].alignment = Alignment(vertical="center", indent=RECUO_LOGO)
    ws.row_dimensions[1].height = 34

    ws["A2"] = subtitulo or ""
    ws["A2"].font = f(size=9, italic=True, color="BFD3E6")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_fim)
    ws["A2"].alignment = Alignment(vertical="center", indent=RECUO_LOGO)
    ws.row_dimensions[2].height = 16


def inserir_logo(ws):
    """Ancora o logo da empresa na faixa institucional, se o arquivo existir."""
    if not os.path.exists(ARQUIVO_LOGO):
        return False
    img = XLImage(ARQUIVO_LOGO)
    alt = 42
    larg = round(img.width * alt / img.height)
    img.width, img.height = larg, alt
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=0, colOff=pixels_to_EMU(8), row=0, rowOff=pixels_to_EMU(5)),
        ext=XDRPositiveSize2D(pixels_to_EMU(larg), pixels_to_EMU(alt)),
    )
    ws.add_image(img)
    return True


def larguras(ws, mapa):
    for col, w in mapa.items():
        ws.column_dimensions[col].width = w


wb = Workbook()

# ==========================================================================
# 1. PARÂMETROS
# ==========================================================================
ws = wb.active
ws.title = "Parâmetros"
ws.sheet_view.showGridLines = False

titulo(ws, "PARÂMETROS DE MARKUP",
       6, "Marv Martins Notari Construtora Ltda — Obra 40200 Empreita Recanto Cataratas Thermas Resort")

ws["A4"] = "FAIXA DE MARKUP SOLICITADA"
ws["A4"].font = f(bold=True, size=11, color=AZUL_ESCURO)

linhas_param = [
    (5, "Markup mínimo (piso da faixa)", 0.35, False,
     "Piso da faixa de 35% a 40% solicitada. Alimenta as colunas '+35%' de todas as abas."),
    (6, "Markup máximo (teto da faixa)", 0.40, False,
     "Teto da faixa de 35% a 40% solicitada. Alimenta as colunas '+40%' de todas as abas."),
    (7, "MARKUP APLICADO", 0.375, True,
     "CÉLULA EDITÁVEL. Altere aqui e TODAS as abas recalculam automaticamente. "
     "Padrão = 37,5% (ponto médio da faixa de 35% a 40%). "
     "Digite 0,35 para 35% ou 0,40 para 40%."),
]
for r, rot, val, destaque, nota in linhas_param:
    ws.cell(row=r, column=1, value=rot).font = f(bold=destaque)
    c = ws.cell(row=r, column=2, value=val)
    c.number_format = PCT
    c.font = f(bold=destaque, color=COR_ENTRADA)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDA
    if destaque:
        c.fill = PatternFill("solid", fgColor=AMARELO)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=AMARELO)
    ws.cell(row=r, column=3, value=nota).font = f(size=9, italic=True, color="595959")

ws["A9"] = "COMO O MARKUP É APLICADO"
ws["A9"].font = f(bold=True, size=11, color=AZUL_ESCURO)
regras = [
    "O markup incide sobre o VALOR UNITÁRIO de mão de obra de cada item.",
    "O total do item é recalculado como: quantidade contratada x valor unitário majorado.",
    "A quantidade contratada NÃO é alterada — apenas o preço.",
    "Como a quantidade é constante, o total majorado equivale ao total original x (1 + markup).",
    "Todos os 6 contratos são de mão de obra: o valor de material é 0,00 em 100% dos itens, "
    "por isso a coluna de material foi omitida.",
    "Itens iniciais e itens de aditivo recebem o mesmo markup.",
]
for i, t in enumerate(regras):
    ws.cell(row=10 + i, column=1, value="•  " + t).font = f(size=10)

r0 = 10 + len(regras) + 1
ws.cell(row=r0, column=1, value="LEGENDA DE CORES").font = f(bold=True, size=11, color=AZUL_ESCURO)
legenda = [
    (COR_ENTRADA, None, "Texto azul",
     "Valor digitado, extraído do contrato em PDF (quantidade e valor unitário)."),
    (COR_FORMULA, None, "Texto preto", "Fórmula calculada na própria aba."),
    (COR_LINK, None, "Texto verde",
     "Link para a aba 'Itens Consolidados', que é a fonte única dos dados."),
    (GRAFITE, AMARELO, "Fundo amarelo", "Célula editável pelo usuário e linhas de total."),
    (COR_FORMULA, AMARELO_SUAVE, "Fundo creme", "Item proveniente de aditivo, não do contrato inicial."),
]
for i, (cor, fundo, nome, desc) in enumerate(legenda):
    r = r0 + 1 + i
    c1 = ws.cell(row=r, column=1, value=nome)
    c1.font = f(bold=True, color=cor)
    c1.alignment = Alignment(horizontal="center")
    c1.border = BORDA
    if fundo:
        c1.fill = PatternFill("solid", fgColor=fundo)
    cd = ws.cell(row=r, column=3, value=desc)
    cd.font = f(size=9, color="595959")

r0 = r0 + len(legenda) + 3
ws.cell(row=r0, column=1, value="ORIGEM DOS DADOS").font = f(bold=True, size=11, color=AZUL_ESCURO)
ws.cell(row=r0 + 1, column=1,
        value="Todos os valores foram extraídos dos contratos em PDF emitidos por JESSICA RIBEIRO "
              "em 21/08/2026, revisão R00. Os arquivos de origem estão indicados no cabeçalho de "
              "cada aba de contrato e na aba 'Conferência'.").font = f(size=9, italic=True, color="595959")
ws.merge_cells(start_row=r0 + 1, start_column=1, end_row=r0 + 1, end_column=6)
ws.cell(row=r0 + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r0 + 1].height = 30

larguras(ws, {"A": 42, "B": 14, "C": 78, "D": 12, "E": 12, "F": 12})

CEL_MIN = "Parâmetros!$B$5"
CEL_MAX = "Parâmetros!$B$6"
CEL_APL = "Parâmetros!$B$7"

# ==========================================================================
# 2. ITENS CONSOLIDADOS  (tabela mestre — fonte única dos dados)
# ==========================================================================
wsi = wb.create_sheet("Itens Consolidados")
wsi.sheet_view.showGridLines = False

COLS_I = [
    ("Contrato", 11), ("Fornecedor", 34), ("Unid. construtiva", 20), ("Origem", 11),
    ("Data do aditivo", 13), ("Referência", 15), ("Código", 9), ("Descrição", 62),
    ("Un.", 7), ("Quantidade", 13), ("Vl. unit. original", 15), ("Total original", 15),
    ("Vl. unit. +35%", 15), ("Total +35%", 15),
    ("Vl. unit. +40%", 15), ("Total +40%", 15),
    ("Vl. unit. aplicado", 15), ("Total aplicado", 15), ("Acréscimo (R$)", 15),
]

titulo(wsi, "ITENS CONSOLIDADOS — TODOS OS CONTRATOS",
       len(COLS_I), "Tabela mestre. Quantidade e valor unitário original (azul) são os únicos "
                    "valores digitados; todo o resto é fórmula. Para incluir novos contratos, "
                    "acrescente linhas ao final desta aba.")

HDR_I = 4
for j, (nome, w) in enumerate(COLS_I, start=1):
    wsi.cell(row=HDR_I, column=j, value=nome)
    wsi.column_dimensions[get_column_letter(j)].width = w
estilo_cabecalho(wsi, HDR_I, 1, len(COLS_I))
wsi.row_dimensions[HDR_I].height = 30

linha = HDR_I + 1
mapa_linhas = {}   # numero do contrato -> lista de (origem, data, linha_planilha)

for c in sorted(CONTRATOS, key=lambda x: x["numero"]):
    mapa_linhas[c["numero"]] = []
    for (origem, data_ad, uc, ref, cod, desc, un, qtd, vl_un) in c["itens"]:
        r = linha
        wsi.cell(row=r, column=1, value=c["numero"])
        wsi.cell(row=r, column=2, value=c["fornecedor"])
        wsi.cell(row=r, column=3, value=uc)
        wsi.cell(row=r, column=4, value=origem)
        wsi.cell(row=r, column=5, value=data_ad)
        wsi.cell(row=r, column=6, value=ref)
        wsi.cell(row=r, column=7, value=cod)
        wsi.cell(row=r, column=8, value=desc)
        wsi.cell(row=r, column=9, value=un)

        cq = wsi.cell(row=r, column=10, value=qtd)
        cq.number_format = QTY
        cq.font = f(color=COR_ENTRADA)

        cv = wsi.cell(row=r, column=11, value=vl_un)
        cv.number_format = CUR4
        cv.font = f(color=COR_ENTRADA)

        wsi.cell(row=r, column=12, value=f"=J{r}*K{r}").number_format = CUR
        wsi.cell(row=r, column=13, value=f"=K{r}*(1+{CEL_MIN})").number_format = CUR4
        wsi.cell(row=r, column=14, value=f"=J{r}*M{r}").number_format = CUR
        wsi.cell(row=r, column=15, value=f"=K{r}*(1+{CEL_MAX})").number_format = CUR4
        wsi.cell(row=r, column=16, value=f"=J{r}*O{r}").number_format = CUR
        wsi.cell(row=r, column=17, value=f"=K{r}*(1+{CEL_APL})").number_format = CUR4
        wsi.cell(row=r, column=18, value=f"=J{r}*Q{r}").number_format = CUR
        wsi.cell(row=r, column=19, value=f"=R{r}-L{r}").number_format = CUR

        for j in range(1, len(COLS_I) + 1):
            cel = wsi.cell(row=r, column=j)
            cel.border = BORDA
            if cel.font.color is None or cel.font.color.rgb in (None, "FF000000"):
                cel.font = f(color=COR_FORMULA)
            cel.alignment = Alignment(vertical="top",
                                      wrap_text=(j == 8),
                                      horizontal="center" if j in (4, 5, 7, 9) else None)
        if origem != "Inicial":
            for j in range(1, len(COLS_I) + 1):
                wsi.cell(row=r, column=j).fill = PatternFill("solid", fgColor=LARANJA_CLARO)

        mapa_linhas[c["numero"]].append((origem, data_ad, r))
        linha += 1

PRIM_I, ULT_I = HDR_I + 1, linha - 1

# Linha de total geral
rt = linha + 1
wsi.cell(row=rt, column=1, value="TOTAL GERAL")
wsi.merge_cells(start_row=rt, start_column=1, end_row=rt, end_column=11)
for col in (12, 14, 16, 18, 19):
    L = get_column_letter(col)
    wsi.cell(row=rt, column=col, value=f"=SUM({L}{PRIM_I}:{L}{ULT_I})").number_format = CUR
for j in range(1, len(COLS_I) + 1):
    cel = wsi.cell(row=rt, column=j)
    cel.font = f(bold=True, size=11, color=GRAFITE)
    cel.fill = PatternFill("solid", fgColor=AMARELO)
    cel.border = BORDA
wsi.cell(row=rt, column=1).alignment = Alignment(horizontal="right", vertical="center")
wsi.row_dimensions[rt].height = 20

wsi.auto_filter.ref = f"A{HDR_I}:S{ULT_I}"
wsi.freeze_panes = f"D{HDR_I + 1}"

SI = "'Itens Consolidados'"

# ==========================================================================
# 3. RESUMO
# ==========================================================================
wsr = wb.create_sheet("Resumo", 1)
wsr.sheet_view.showGridLines = False

COLS_R = [
    ("Contrato", 11), ("Fornecedor", 36), ("CNPJ", 20), ("Objeto do contrato", 52),
    ("Início", 11), ("Término", 11), ("Nº itens", 9),
    ("Valor inicial", 15), ("Valor aditivos", 15), ("Valor total original", 17),
    ("Total +35%", 16), ("Total +40%", 16), ("Total aplicado", 16), ("Acréscimo (R$)", 15),
]

titulo(wsr, "RESUMO DOS CONTRATOS COM MARKUP",
       len(COLS_R), "Todos os valores calculados a partir da aba 'Itens Consolidados'. "
                    "O percentual aplicado é o da célula Parâmetros!B7.")

wsr["A3"] = "Markup aplicado:"
wsr["A3"].font = f(bold=True, size=10)
wsr["A3"].alignment = Alignment(horizontal="right", vertical="center")
wsr.merge_cells("A3:B3")
wsr["C3"] = f"={CEL_APL}"
wsr["C3"].number_format = PCT
wsr["C3"].font = f(bold=True, size=11, color=COR_LINK)
wsr["C3"].fill = PatternFill("solid", fgColor=AMARELO)
wsr["C3"].alignment = Alignment(horizontal="center")
wsr["C3"].border = BORDA
wsr["D3"] = "← altere em Parâmetros!B7 para recalcular a planilha inteira"
wsr["D3"].font = f(size=9, italic=True, color="595959")

HDR_R = 5
for j, (nome, w) in enumerate(COLS_R, start=1):
    wsr.cell(row=HDR_R, column=j, value=nome)
    wsr.column_dimensions[get_column_letter(j)].width = w
estilo_cabecalho(wsr, HDR_R, 1, len(COLS_R))
wsr.row_dimensions[HDR_R].height = 30

RG_CT = f"{SI}!$A${PRIM_I}:$A${ULT_I}"
RG_OR = f"{SI}!$D${PRIM_I}:$D${ULT_I}"


def rg(col):
    return f"{SI}!${col}${PRIM_I}:${col}${ULT_I}"


r = HDR_R + 1
prim_r = r
for c in sorted(CONTRATOS, key=lambda x: x["numero"]):
    wsr.cell(row=r, column=1, value=c["numero"])
    wsr.cell(row=r, column=2, value=c["fornecedor"])
    wsr.cell(row=r, column=3, value=c["cnpj"])
    wsr.cell(row=r, column=4, value=c["objeto"])
    wsr.cell(row=r, column=5, value=c["inicio"])
    wsr.cell(row=r, column=6, value=c["termino"])
    wsr.cell(row=r, column=7, value=f"=COUNTIFS({RG_CT},$A{r})")
    wsr.cell(row=r, column=8, value=f'=SUMIFS({rg("L")},{RG_CT},$A{r},{RG_OR},"Inicial")')
    wsr.cell(row=r, column=9, value=f"=J{r}-H{r}")
    wsr.cell(row=r, column=10, value=f'=SUMIFS({rg("L")},{RG_CT},$A{r})')
    wsr.cell(row=r, column=11, value=f'=SUMIFS({rg("N")},{RG_CT},$A{r})')
    wsr.cell(row=r, column=12, value=f'=SUMIFS({rg("P")},{RG_CT},$A{r})')
    wsr.cell(row=r, column=13, value=f'=SUMIFS({rg("R")},{RG_CT},$A{r})')
    wsr.cell(row=r, column=14, value=f"=M{r}-J{r}")

    for j in range(1, len(COLS_R) + 1):
        cel = wsr.cell(row=r, column=j)
        cel.border = BORDA
        cel.font = f(color=COR_LINK if j >= 7 else COR_FORMULA)
        cel.alignment = Alignment(vertical="top", wrap_text=(j == 4),
                                  horizontal="center" if j in (1, 5, 6, 7) else None)
        if j >= 8:
            cel.number_format = CUR
    wsr.cell(row=r, column=1).font = f(bold=True, color=COR_FORMULA)
    wsr.cell(row=r, column=7).number_format = "0"
    r += 1

ult_r = r - 1
for col in (8, 9, 10, 11, 12, 13, 14):
    L = get_column_letter(col)
    wsr.cell(row=r, column=col, value=f"=SUM({L}{prim_r}:{L}{ult_r})").number_format = CUR
wsr.cell(row=r, column=7, value=f"=SUM(G{prim_r}:G{ult_r})").number_format = "0"
wsr.cell(row=r, column=1, value="TOTAL GERAL")
wsr.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
for j in range(1, len(COLS_R) + 1):
    cel = wsr.cell(row=r, column=j)
    cel.font = f(bold=True, size=11, color=GRAFITE)
    cel.fill = PatternFill("solid", fgColor=AMARELO)
    cel.border = BORDA
    cel.alignment = Alignment(horizontal="center", vertical="center")
wsr.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
wsr.row_dimensions[r].height = 20
LINHA_TOT_RESUMO = r

wsr.freeze_panes = f"A{HDR_R + 1}"

# Bloco comparativo dos três cenários
rb = r + 3
wsr.cell(row=rb, column=1, value="COMPARATIVO DE CENÁRIOS").font = f(bold=True, size=12, color=AZUL_ESCURO)
wsr.cell(row=rb + 1, column=1, value="Cenário")
wsr.merge_cells(start_row=rb + 1, start_column=1, end_row=rb + 1, end_column=2)
wsr.cell(row=rb + 1, column=3, value="Markup")
wsr.cell(row=rb + 1, column=4, value="Total dos 6 contratos")
wsr.cell(row=rb + 1, column=5, value="Acréscimo sobre o original")
estilo_cabecalho(wsr, rb + 1, 1, 5)

cenarios = [
    ("Valor original (sem markup)", None, f"=J{LINHA_TOT_RESUMO}", "0"),
    ("Piso da faixa", CEL_MIN, f"=K{LINHA_TOT_RESUMO}", None),
    ("Teto da faixa", CEL_MAX, f"=L{LINHA_TOT_RESUMO}", None),
    ("APLICADO NA PLANILHA", CEL_APL, f"=M{LINHA_TOT_RESUMO}", None),
]
for i, (nome, cel_pct, formula, _) in enumerate(cenarios):
    rr = rb + 2 + i
    destaque = nome.startswith("APLICADO")
    wsr.cell(row=rr, column=1, value=nome).font = f(bold=destaque)
    wsr.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
    cp = wsr.cell(row=rr, column=3, value=(f"={cel_pct}" if cel_pct else 0))
    cp.number_format = PCT
    cp.font = f(bold=destaque, color=COR_LINK if cel_pct else COR_FORMULA)
    cp.alignment = Alignment(horizontal="center")
    cv = wsr.cell(row=rr, column=4, value=formula)
    cv.number_format = CUR
    cv.font = f(bold=destaque, color=COR_FORMULA)
    ca = wsr.cell(row=rr, column=5, value=f"=D{rr}-$J${LINHA_TOT_RESUMO}")
    ca.number_format = CUR
    ca.font = f(bold=destaque, color=COR_FORMULA)
    for j in range(1, 6):
        wsr.cell(row=rr, column=j).border = BORDA
        if destaque:
            wsr.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=AMARELO)

# ==========================================================================
# 4. UMA ABA POR CONTRATO
# ==========================================================================
COLS_C = [
    ("Origem", 11), ("Data do aditivo", 13), ("Unid. construtiva", 20), ("Referência", 15),
    ("Código", 9), ("Descrição", 62), ("Un.", 7), ("Quantidade", 13),
    ("Vl. unit. original", 15), ("Total original", 15),
    ("Vl. unit. +35%", 15), ("Total +35%", 15),
    ("Vl. unit. +40%", 15), ("Total +40%", 15),
    ("Vl. unit. aplicado", 15), ("Total aplicado", 15),
]
NC = len(COLS_C)

# coluna da aba de contrato -> coluna correspondente em 'Itens Consolidados'
MAPA_COL = {1: "D", 2: "E", 3: "C", 4: "F", 5: "G", 6: "H", 7: "I", 8: "J",
            9: "K", 10: "L", 11: "M", 12: "N", 13: "O", 14: "P", 15: "Q", 16: "R"}

linhas_totais_contrato = {}

for c in sorted(CONTRATOS, key=lambda x: x["numero"]):
    wsc = wb.create_sheet(c["aba"])
    wsc.sheet_view.showGridLines = False

    titulo(wsc, f'{c["numero"]} — {c["fornecedor"]}', NC,
           f'Arquivo de origem: {c["arquivo"]}  |  Obra: {c["obra"]}')

    info = [
        ("CNPJ do fornecedor", c["cnpj"]),
        ("Data do contrato", c["data_contrato"]),
        ("Vigência", f'{c["inicio"]} a {c["termino"]}'),
        ("Tipo de contrato", c["tipo"]),
        ("Objeto", c["objeto"]),
        ("Impostos / retenção", c["impostos"]),
    ]
    r = 4
    for rot, val in info:
        wsc.cell(row=r, column=1, value=rot).font = f(bold=True, size=9)
        for j in range(1, 4):
            wsc.cell(row=r, column=j).fill = PatternFill("solid", fgColor=CINZA_CLARO)
            wsc.cell(row=r, column=j).border = BORDA
        wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        wsc.cell(row=r, column=1).alignment = Alignment(vertical="center")
        cel = wsc.cell(row=r, column=4, value=val)
        cel.font = f(size=9)
        cel.alignment = Alignment(vertical="center", wrap_text=True)
        wsc.merge_cells(start_row=r, start_column=4, end_row=r, end_column=NC)
        r += 1

    if c["aditivos_desc"]:
        r += 1
        wsc.cell(row=r, column=1, value="ADITIVOS").font = f(bold=True, size=11, color=AZUL_ESCURO)
        r += 1
        for nome_ad, desc_ad in c["aditivos_desc"].items():
            data_ad = next((d for o, d, _ in mapa_linhas[c["numero"]] if o == nome_ad), "")
            wsc.cell(row=r, column=1, value=f"{nome_ad} — {data_ad}").font = f(bold=True, size=9)
            for j in range(1, 4):
                wsc.cell(row=r, column=j).fill = PatternFill("solid", fgColor=LARANJA_CLARO)
                wsc.cell(row=r, column=j).border = BORDA
            wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            wsc.cell(row=r, column=1).alignment = Alignment(vertical="top")
            cel = wsc.cell(row=r, column=4, value=desc_ad)
            cel.font = f(size=9)
            cel.alignment = Alignment(vertical="top", wrap_text=True)
            wsc.merge_cells(start_row=r, start_column=4, end_row=r, end_column=NC)
            wsc.row_dimensions[r].height = max(14, 12 * (len(desc_ad) // 215 + 1))
            r += 1

    r += 1
    hdr = r
    for j, (nome, w) in enumerate(COLS_C, start=1):
        wsc.cell(row=hdr, column=j, value=nome)
        wsc.column_dimensions[get_column_letter(j)].width = w
    estilo_cabecalho(wsc, hdr, 1, NC)
    wsc.row_dimensions[hdr].height = 30

    r = hdr + 1
    prim_c = r
    linhas_subtotal = []

    # agrupa por origem, preservando a ordem (Inicial, Aditivo 1, 2, 3)
    ordem = []
    for origem, _, _ in mapa_linhas[c["numero"]]:
        if origem not in ordem:
            ordem.append(origem)

    for origem in ordem:
        refs = [rl for o, _, rl in mapa_linhas[c["numero"]] if o == origem]
        ini_grupo = r
        for rl in refs:
            for j in range(1, NC + 1):
                col_i = MAPA_COL[j]
                ref = f"{SI}!{col_i}{rl}"
                # colunas de texto podem estar vazias na origem (ex.: data de aditivo
                # em item inicial); sem o IF o link devolveria 0 em vez de vazio
                formula = f'=IF({ref}="","",{ref})' if j <= 7 else f"={ref}"
                cel = wsc.cell(row=r, column=j, value=formula)
                cel.border = BORDA
                cel.font = f(color=COR_LINK)
                cel.alignment = Alignment(vertical="top", wrap_text=(j == 6),
                                          horizontal="center" if j in (1, 2, 5, 7) else None)
                if j == 8:
                    cel.number_format = QTY
                elif j in (9, 11, 13, 15):
                    cel.number_format = CUR4
                elif j in (10, 12, 14, 16):
                    cel.number_format = CUR
            if origem != "Inicial":
                for j in range(1, NC + 1):
                    wsc.cell(row=r, column=j).fill = PatternFill("solid", fgColor=LARANJA_CLARO)
            r += 1

        # subtotal do grupo
        rot = "Total dos itens iniciais" if origem == "Inicial" else f"Total do {origem.lower()}"
        wsc.cell(row=r, column=1, value=rot)
        wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        for col in (10, 12, 14, 16):
            L = get_column_letter(col)
            wsc.cell(row=r, column=col, value=f"=SUM({L}{ini_grupo}:{L}{r - 1})").number_format = CUR
        for j in range(1, NC + 1):
            cel = wsc.cell(row=r, column=j)
            cel.font = f(bold=True, size=10)
            cel.fill = PatternFill("solid", fgColor=CINZA_MEDIO)
            cel.border = BORDA
        wsc.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
        linhas_subtotal.append(r)
        r += 1

    # total do contrato
    wsc.cell(row=r, column=1, value=f'TOTAL DO CONTRATO {c["numero"]}')
    wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    for col in (10, 12, 14, 16):
        L = get_column_letter(col)
        parcelas = "+".join(f"{L}{s}" for s in linhas_subtotal)
        wsc.cell(row=r, column=col, value=f"={parcelas}").number_format = CUR
    for j in range(1, NC + 1):
        cel = wsc.cell(row=r, column=j)
        cel.font = f(bold=True, size=11, color=GRAFITE)
        cel.fill = PatternFill("solid", fgColor=AMARELO)
        cel.border = BORDA
    wsc.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
    wsc.row_dimensions[r].height = 20
    linhas_totais_contrato[c["numero"]] = (c["aba"], r)

    # acréscimo gerado pelo markup
    r += 1
    wsc.cell(row=r, column=1, value="Acréscimo gerado pelo markup aplicado")
    wsc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    wsc.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
    wsc.cell(row=r, column=1).font = f(bold=True, italic=True, size=10)
    cel = wsc.cell(row=r, column=16, value=f"=P{r - 1}-J{r - 1}")
    cel.number_format = CUR
    cel.font = f(bold=True, size=10)
    cel.fill = PatternFill("solid", fgColor=VERDE_CLARO)
    cel.border = BORDA

    wsc.freeze_panes = f"A{hdr + 1}"

# ==========================================================================
# 5. CONFERÊNCIA
# ==========================================================================
wsf = wb.create_sheet("Conferência")
wsf.sheet_view.showGridLines = False

titulo(wsf, "CONFERÊNCIA — CALCULADO x IMPRESSO NO PDF", 7,
       "Confronta a soma dos itens digitados nesta planilha com os valores de contrato "
       "impressos em cada PDF de origem. Todas as linhas devem estar OK.")

cab = ["Contrato", "Arquivo de origem (PDF)", "Item conferido",
       "Calculado na planilha", "Impresso no PDF", "Diferença", "Status"]
larg = [11, 44, 20, 19, 17, 14, 11]
HDR_F = 4
for j, (nome, w) in enumerate(zip(cab, larg), start=1):
    wsf.cell(row=HDR_F, column=j, value=nome)
    wsf.column_dimensions[get_column_letter(j)].width = w
estilo_cabecalho(wsf, HDR_F, 1, 7)
wsf.row_dimensions[HDR_F].height = 28

r = HDR_F + 1
prim_f = r
for c in sorted(CONTRATOS, key=lambda x: x["numero"]):
    lr = next(i for i in range(HDR_R + 1, ult_r + 1) if wsr.cell(row=i, column=1).value == c["numero"])
    for rotulo, col_resumo, valor_pdf in (
        ("Valor inicial", "H", c["conf_inicial"]),
        ("Valor de aditivos", "I", c["conf_aditivos"]),
        ("Valor do contrato", "J", c["conf_total"]),
    ):
        wsf.cell(row=r, column=1, value=c["numero"])
        wsf.cell(row=r, column=2, value=c["arquivo"])
        wsf.cell(row=r, column=3, value=rotulo)
        wsf.cell(row=r, column=4, value=f"=Resumo!{col_resumo}{lr}").number_format = CUR
        cp = wsf.cell(row=r, column=5, value=valor_pdf)
        cp.number_format = CUR
        cp.font = f(color=COR_ENTRADA)
        wsf.cell(row=r, column=6, value=f"=D{r}-E{r}").number_format = CUR
        wsf.cell(row=r, column=7, value=f'=IF(ROUND(F{r},2)=0,"OK","DIVERGE")')
        for j in range(1, 8):
            cel = wsf.cell(row=r, column=j)
            cel.border = BORDA
            if j != 5:
                cel.font = f(color=COR_LINK if j == 4 else COR_FORMULA,
                             bold=(j == 7))
            cel.alignment = Alignment(horizontal="center" if j in (1, 3, 7) else None,
                                      vertical="center")
        if rotulo == "Valor do contrato":
            for j in range(1, 8):
                wsf.cell(row=r, column=j).fill = PatternFill("solid", fgColor=CINZA_CLARO)
        r += 1

wsf.cell(row=r, column=5).comment = Comment(
    "Valores impressos no campo 'Informações do contrato' de cada PDF "
    "(Valor do inicial / Valor de aditivos / Valor do contrato), emitidos por "
    "JESSICA RIBEIRO em 21/08/2026, revisão R00.", "Conferência", height=110, width=320)

ult_f = r - 1
wsf.cell(row=r + 1, column=3, value="RESULTADO DA CONFERÊNCIA").font = f(bold=True, size=11)
cel = wsf.cell(row=r + 1, column=7,
               value=f'=IF(COUNTIF(G{prim_f}:G{ult_f},"DIVERGE")=0,"TUDO OK","VERIFICAR")')
cel.font = f(bold=True, size=11, color="FFFFFF")
cel.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
cel.border = BORDA
cel.alignment = Alignment(horizontal="center")

wsf.cell(row=r + 3, column=1,
         value="Observação: o arquivo CT196__Fabio_Gesso__R00.pdf traz como razão social do "
               "fornecedor ROSENILDE RODRIGUES LINDOLFO DE SOUZA (CNPJ 42.578.638/0001-54). "
               "O nome do arquivo e a razão social do contrato não coincidem — confirmar qual "
               "deve constar no cadastro.").font = f(size=9, italic=True, color="C00000")
wsf.merge_cells(start_row=r + 3, start_column=1, end_row=r + 4, end_column=7)
wsf.cell(row=r + 3, column=1).alignment = Alignment(wrap_text=True, vertical="top")

# ==========================================================================
# 6. AJUSTES FINAIS — impressão e cores das abas
# ==========================================================================
CORES_ABA = {"Parâmetros": AMARELO, "Resumo": GRAFITE,
             "Itens Consolidados": GRAFITE, "Conferência": "15803D"}

logos = 0
for aba in wb.worksheets:
    logos += inserir_logo(aba)
    aba.sheet_properties.tabColor = CORES_ABA.get(aba.title, CINZA_LINHA)
    aba.page_setup.orientation = "landscape"
    aba.page_setup.paperSize = aba.PAPERSIZE_A4
    aba.page_setup.fitToWidth = 1
    aba.page_setup.fitToHeight = 0
    aba.sheet_properties.pageSetUpPr.fitToPage = True
    aba.print_options.horizontalCentered = True
    aba.page_margins.left = aba.page_margins.right = 0.3
    aba.page_margins.top = aba.page_margins.bottom = 0.4
    aba.oddFooter.right.text = "Página &P de &N"
    aba.oddFooter.right.size = 8
    aba.oddFooter.left.text = aba.title
    aba.oddFooter.left.size = 8

# repete a linha de cabeçalho em todas as páginas impressas das tabelas longas
wsi.print_title_rows = f"{HDR_I}:{HDR_I}"
wsr.print_title_rows = f"{HDR_R}:{HDR_R}"
wsf.print_title_rows = f"{HDR_F}:{HDR_F}"

# ordem final das abas
wb.move_sheet("Parâmetros", offset=-wb.sheetnames.index("Parâmetros"))
wb._sheets.sort(key=lambda s: (
    {"Parâmetros": 0, "Resumo": 1, "Itens Consolidados": 2}.get(s.title, 3),
    s.title if s.title.startswith("CT-") else ("zz" if s.title == "Conferência" else s.title),
))

wb.save(ARQUIVO_SAIDA)
print(f"Gerado: {ARQUIVO_SAIDA}")
print("Abas:", wb.sheetnames)
print(f"Itens na tabela mestre: {ULT_I - PRIM_I + 1} (linhas {PRIM_I} a {ULT_I})")
if logos:
    print(f"Logo '{ARQUIVO_LOGO}' embutido em {logos} abas.")
else:
    print(f"AVISO: '{ARQUIVO_LOGO}' não encontrado — faixa gerada sem o logo. "
          f"Salve o logo oficial como '{ARQUIVO_LOGO}' nesta pasta e rode de novo.")
